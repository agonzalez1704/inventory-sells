# Seed compatibility tags for Ruli from GROB's human-readable ACES xlsx files
# (PartNumber -> Marca/Modelo/SubModelo/Años). One tag per vehicle+year-range,
# e.g. "Acura RDX 2007-2012"; per-year rows collapse into one range.
#
# Idempotent: tags dedupe on nombre_norm, links on (product_id, tag_id).
# Re-run whenever GROB ships a new applications file (add its path below).
#
#   python3 scripts/sembrar-tags-aces.py           # dry run
#   python3 scripts/sembrar-tags-aces.py --sembrar

import json
import sys
import unicodedata
import urllib.request
from pathlib import Path

import openpyxl

# The full fitment lives in "Catálogo explosionado" (one row per part-vehicle-
# year, ~393k rows); the second file only adds brand-new part numbers.
EXPLOSIONADO = "/Users/antoniogonzalez/MEGA/ACES & PIES agosto/In House GROB (Archivo para ACES) correcciones Julio 2026.xlsx"
ARCHIVOS = [
    "/Users/antoniogonzalez/MEGA/ACES & PIES agosto/Nuevos numeros de parte agosto 2026.xlsx",
]
SEMBRAR = "--sembrar" in sys.argv

creds = json.loads((Path(__file__).parent.parent / ".insforge/negocios.json").read_text())["ruli"]


def rawsql(query):
    req = urllib.request.Request(
        f"{creds['oss_host']}/api/database/advance/rawsql/unrestricted",
        data=json.dumps({"query": query}).encode(),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {creds['api_key']}"},
        method="POST",
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read()).get("rows", [])


def norm(s):
    s = unicodedata.normalize("NFD", str(s).lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()


def celdas(row):
    return ["" if c is None else str(c).strip() for c in row]


def leer(path):
    """Yield (sku, marca, modelo, sub, anio_ini, anio_fin) rows from any sheet
    whose header row carries Marca+Modelo and a part-number column."""
    wb = openpyxl.load_workbook(path, read_only=True)
    for ws in wb.worksheets:
        header = None
        idx = {}
        for row in ws.iter_rows(values_only=True):
            vals = celdas(row)
            if header is None:
                lower = [v.lower() for v in vals]
                if "marca" in lower and "modelo" in lower:
                    header = lower
                    idx = {name: i for i, name in enumerate(lower)}
                continue
            def col(*names):
                for n in names:
                    if n in idx and idx[n] < len(vals):
                        return vals[idx[n]]
                return ""
            sku = col("partnumber", "codigo", "código")
            marca, modelo = col("marca"), col("modelo")
            if not sku or not marca or not modelo or modelo == "#":
                continue
            sub = col("sub modelo", "sub", "submodelo")
            if sub == "#":
                sub = ""
            ini = col("inicial", "year")
            fin = col("final") or ini
            yield sku, marca, modelo, sub, ini, fin


def leer_explosionado():
    """The big sheet: fixed columns, corrected values preferred over raw ones,
    '#' meaning empty throughout."""
    wb = openpyxl.load_workbook(EXPLOSIONADO, read_only=True)
    ws = wb["Catálogo explosionado"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    for row in rows:
        v = celdas(row)
        if len(v) < 28:
            continue
        def pick(corr, raw):
            x = v[corr] if v[corr] not in ("", "#", "-") else v[raw]
            return "" if x in ("", "#", "-") else x
        sku = v[4]
        marca = pick(10, 9)
        modelo = pick(12, 11)
        if not sku or not marca or not modelo:
            continue
        sub = pick(14, 13)
        ini = pick(25, 23) or v[27]
        fin = pick(26, 24) or v[27]
        yield sku, marca, modelo, sub, ini, fin


def main():
    # (sku, vehiculo) -> [min_year, max_year]; per-year rows collapse here.
    rangos = {}
    fuentes = [leer_explosionado()]
    for path in ARCHIVOS:
        fuentes.append(leer(path))
    for fuente in fuentes:
        for sku, marca, modelo, sub, ini, fin in fuente:
            vehiculo = " ".join(x for x in (marca, modelo, sub) if x)
            key = (sku, vehiculo)
            try:
                a, b = int(float(ini)), int(float(fin))
            except ValueError:
                continue
            r = rangos.setdefault(key, [a, b])
            r[0], r[1] = min(r[0], a), max(r[1], b)

    print(f"aplicaciones (sku, vehiculo): {len(rangos)}")

    skus = {s.lower() for s, _ in rangos}
    filas = rawsql("SELECT id, sku FROM products")
    por_sku = {r["sku"].lower(): r["id"] for r in filas}
    pares = []  # (product_id, tag_nombre)
    for (sku, vehiculo), (a, b) in rangos.items():
        pid = por_sku.get(sku.lower())
        if not pid:
            continue
        nombre = f"{vehiculo} {a}" if a == b else f"{vehiculo} {a}-{b}"
        pares.append((pid, nombre))
    print(f"con producto en catálogo: {len(pares)}")
    print(f"etiquetas distintas: {len({n for _, n in pares})}")
    if not SEMBRAR:
        print("\nDry run. Corre con --sembrar para escribir.")
        return

    def sq(s):
        return s.replace("'", "''")

    nombres = sorted({n for _, n in pares})
    for i in range(0, len(nombres), 300):
        vals = ",".join(f"('{sq(n)}', '{sq(norm(n))}')" for n in nombres[i : i + 300])
        rawsql(
            f"INSERT INTO tags (nombre, nombre_norm) VALUES {vals} "
            "ON CONFLICT (nombre_norm) DO NOTHING"
        )
    for i in range(0, len(pares), 300):
        vals = ",".join(
            f"('{pid}'::uuid, '{sq(norm(n))}')" for pid, n in pares[i : i + 300]
        )
        rawsql(
            "INSERT INTO product_tags (product_id, tag_id) "
            f"SELECT v.pid, t.id FROM (VALUES {vals}) AS v(pid, nn) "
            "JOIN tags t ON t.nombre_norm = v.nn "
            "ON CONFLICT DO NOTHING"
        )
    tot = rawsql("SELECT count(*) AS c FROM product_tags")[0]["c"]
    print(f"✓ sembrado. product_tags totales: {tot}")


main()
